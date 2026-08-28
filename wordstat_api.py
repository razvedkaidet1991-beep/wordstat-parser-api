#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Парсер Яндекс Вордстат через Cloud Search API (topRequests).
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import time
from typing import Any
from urllib.parse import urlencode

import openpyxl
import requests
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(BASE_DIR, "input")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
DATA_DIR = os.path.join(BASE_DIR, "data")

QUERIES_FILE = os.path.join(INPUT_DIR, "queries.txt")
STOP_WORDS_FILE = os.path.join(INPUT_DIR, "stop_words.txt")
CREDENTIALS_FILE = os.path.join(DATA_DIR, "credentials.txt")
WORDS_OUTPUT_FILE = os.path.join(OUTPUT_DIR, "wordstat_words.txt")
REPORT_OUTPUT_FILE = os.path.join(OUTPUT_DIR, "wordstat_report.xlsx")
INVALID_QUERIES_FILE = os.path.join(OUTPUT_DIR, "invalid_queries.txt")

API_URL = "https://searchapi.api.cloud.yandex.net/v2/wordstat/topRequests"
REQUEST_DELAY_SEC = 0.25
MAX_PHRASES = 2000
MAX_CONSECUTIVE_429 = 10
WORDSTAT_WEB = "https://wordstat.yandex.ru/"

EXCEL_HEADERS = [
    "Запрос",
    "Частота (базовая)",
    "Частота (точная)",
    "Частота (уточненная)",
]

FREQ_TYPES_ORDER = ("base", "exact", "precise")
FREQ_TYPE_ALIASES = {
    "b": "base",
    "base": "base",
    "базовая": "base",
    "e": "exact",
    "exact": "exact",
    "точная": "exact",
    "p": "precise",
    "precise": "precise",
    "уточнённая": "precise",
    "уточненная": "precise",
}
FREQ_TYPE_LABELS = {
    "base": "базовая",
    "exact": "точная",
    "precise": "уточнённая",
}
FREQ_TYPE_ROW_KEYS = {
    "base": "base_frequency",
    "exact": "exact_frequency",
    "precise": "precise_frequency",
}
DEFAULT_FREQ_TYPES: frozenset[str] = frozenset({"base"})


class RateLimitError(RuntimeError):
    """HTTP 429 — превышена квота / rate limit."""


class QuotaExceededStop(RuntimeError):
    """Остановка после MAX_CONSECUTIVE_429 подряд ошибок 429."""


class InvalidQueryError(RuntimeError):
    """HTTP 400 Invalid query."""


def ensure_dirs() -> None:
    for path in (INPUT_DIR, OUTPUT_DIR, DATA_DIR):
        os.makedirs(path, exist_ok=True)


def load_credentials(path: str = CREDENTIALS_FILE) -> tuple[str, str]:
    """Читает api_key и folder_id из data/credentials.txt."""
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Нет файла {path}. Скопируйте data/credentials.example.txt "
            f"в data/credentials.txt и заполните ключи. См. docs/API_KEY.md"
        )

    api_key = ""
    folder_id = ""
    with open(path, "r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            if "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip().lower()
            value = value.strip().strip('"').strip("'")
            if key in ("api_key", "apikey", "secret", "key"):
                api_key = value
            elif key in ("folder_id", "folderid", "folder"):
                folder_id = value

    if not api_key or not folder_id:
        raise ValueError(
            f"В {path} нужны строки api_key=... и folder_id=... "
            f"(см. data/credentials.example.txt)"
        )
    return api_key, folder_id


def read_lines(path: str) -> list[str]:
    if not os.path.exists(path):
        return []
    with open(path, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]


def write_queries_file(remaining: list[str], path: str = QUERIES_FILE) -> None:
    """Перезаписывает queries.txt оставшимися запросами."""
    ensure_dirs()
    with open(path, "w", encoding="utf-8") as f:
        for line in remaining:
            f.write(line + "\n")


def append_invalid_query(
    seed: str,
    normalized: str,
    reason: str,
    path: str = INVALID_QUERIES_FILE,
) -> None:
    """Логирует невалидные seed-запросы в output/invalid_queries.txt."""
    ensure_dirs()
    with open(path, "a", encoding="utf-8") as f:
        f.write(f"{seed}\t{normalized}\t{reason}\n")


def normalize_query_for_api(query: str) -> str:
    """
    Нормализует seed-запрос для API:
    - кавычки всех видов и ! убираются (операторы навешивает format_query)
    - &, /, |, @, ;, : заменяются пробелами
    - сохраняются буквы/цифры/пробел и символы + -
    - схлопываются лишние пробелы
    """
    q = (query or "").strip()
    if not q:
        return ""
    # Кавычки и ! — не часть текста seed; "..." / "! !" добавляет format_query
    q = re.sub(r'[\"\'«»„“”`!]', " ", q)
    q = re.sub(r"[&/|@;:]", " ", q)
    q = re.sub(r"[^\w\s+\-]", " ", q, flags=re.UNICODE)
    q = re.sub(r"\s+", " ", q, flags=re.UNICODE).strip()
    return q


def format_query(query: str, query_type: str = "base") -> str:
    query = query.strip()
    if query_type == "exact":
        return f'"{query}"'
    if query_type == "precise":
        words = query.split()
        return '"' + " ".join(f"!{w}" for w in words) + '"'
    return query


def parse_freq_types(raw: str | None) -> set[str]:
    """
    Парсит мультивыбор типов частот.
    Примеры: 'b', 'be', 'bep', 'base,exact', 'e p'.
    Пустая строка / None → только базовая.
    """
    if raw is None:
        return set(DEFAULT_FREQ_TYPES)

    text = raw.strip().lower()
    if not text:
        return set(DEFAULT_FREQ_TYPES)

    tokens: list[str] = []
    if re.fullmatch(r"[bep]+", text):
        tokens = list(text)
    else:
        tokens = [t for t in re.split(r"[\s,;+|]+", text) if t]

    selected: set[str] = set()
    unknown: list[str] = []
    for token in tokens:
        mapped = FREQ_TYPE_ALIASES.get(token)
        if mapped is None:
            unknown.append(token)
            continue
        selected.add(mapped)

    if unknown:
        raise ValueError(
            f"Неизвестные типы частот: {', '.join(unknown)}. "
            f"Допустимо: b/e/p или base/exact/precise"
        )
    if not selected:
        return set(DEFAULT_FREQ_TYPES)
    return selected


def format_freq_types(freq_types: set[str]) -> str:
    labels = [FREQ_TYPE_LABELS[t] for t in FREQ_TYPES_ORDER if t in freq_types]
    return ", ".join(labels)


def build_wordstat_url(query: str) -> str:
    params = {"region": "all", "view": "table", "words": query}
    return f"{WORDSTAT_WEB}?{urlencode(params)}"


class WordstatAPI:
    def __init__(self, api_key: str, folder_id: str, delay: float = REQUEST_DELAY_SEC):
        self.api_key = api_key
        self.folder_id = folder_id
        self.delay = delay
        self.consecutive_429 = 0
        self.session = requests.Session()
        self.session.headers.update(
            {
                "Authorization": f"Api-Key {api_key}",
                "Content-Type": "application/json; charset=utf-8",
            }
        )

    def _note_success(self) -> None:
        self.consecutive_429 = 0

    def _note_429(self) -> None:
        self.consecutive_429 += 1
        print(f"  ⚠ 429 ({self.consecutive_429}/{MAX_CONSECUTIVE_429})")
        if self.consecutive_429 >= MAX_CONSECUTIVE_429:
            raise QuotaExceededStop(
                f"Остановка: {MAX_CONSECUTIVE_429} ответов 429 подряд. "
                f"Квота исчерпана — повторите позже."
            )

    def top_requests(self, phrase: str, num_phrases: int = 5) -> dict[str, Any]:
        body = {
            "phrase": phrase,
            "numPhrases": max(1, min(int(num_phrases), MAX_PHRASES)),
            "folderId": self.folder_id,
        }
        try:
            resp = self.session.post(API_URL, json=body, timeout=60)
        except requests.RequestException as e:
            raise RuntimeError(f"Сеть / запрос не выполнен: {e}") from e

        if resp.status_code == 401:
            raise RuntimeError("401 Unauthorized — проверьте api_key в data/credentials.txt")
        if resp.status_code == 403:
            raise RuntimeError(
                "403 Forbidden — проверьте роль search-api.webSearch.user, "
                "scope yc.search-api.execute и биллинг (docs/API_KEY.md)"
            )
        if resp.status_code == 429:
            self._note_429()
            raise RateLimitError("429 Too Many Requests — превышена квота / rate limit")
        if resp.status_code == 400:
            message = resp.text[:500]
            if "Invalid query" in message:
                raise InvalidQueryError(f"HTTP 400 Invalid query: {message}")
            raise RuntimeError(f"HTTP 400: {message}")
        if resp.status_code >= 400:
            raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:500]}")

        self._note_success()
        data = resp.json()
        time.sleep(self.delay)
        return data

    def get_total_count(self, phrase: str) -> int | None:
        data = self.top_requests(phrase, num_phrases=1)
        total = data.get("totalCount")
        if total is None:
            return None
        try:
            return int(str(total).replace(" ", "").replace("\xa0", ""))
        except ValueError:
            return None

    def get_popular_phrases(self, phrase: str, num_phrases: int = MAX_PHRASES) -> list[str]:
        data = self.top_requests(phrase, num_phrases=num_phrases)
        results = data.get("results") or []
        phrases: list[str] = []
        seen: set[str] = set()
        for item in results:
            text = (item.get("phrase") or "").strip()
            if not text:
                continue
            key = text.lower()
            if key in seen:
                continue
            seen.add(key)
            phrases.append(text)
        return phrases


def filter_stop_words(
    phrases: list[str], stop_words: list[str], *, quiet: bool = False
) -> list[str]:
    if not stop_words:
        if not quiet:
            print("ℹ️  Стоп-слова не используются")
        return list(phrases)

    stops = [s.lower() for s in stop_words if s]
    filtered: list[str] = []
    removed = 0
    for phrase in phrases:
        lower = phrase.lower()
        if any(stop in lower for stop in stops):
            removed += 1
            continue
        filtered.append(phrase)
    if not quiet:
        print(f"  фильтр стоп-слов: −{removed}, прошло {len(filtered)}")
    return filtered


def append_phrases(phrases: list[str], path: str = WORDS_OUTPUT_FILE) -> int:
    """Дописывает фразы в конец файла. Возвращает число записанных строк."""
    if not phrases:
        return 0
    ensure_dirs()
    with open(path, "a", encoding="utf-8") as f:
        for phrase in phrases:
            f.write(phrase + "\n")
    return len(phrases)


def append_excel_row(row: dict, path: str = REPORT_OUTPUT_FILE) -> None:
    """Добавляет одну строку результата в Excel (создаёт файл при необходимости)."""
    ensure_dirs()
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Анализ запросов Вордстат"
        for col, header in enumerate(EXCEL_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

    next_row = ws.max_row + 1
    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        for col, header in enumerate(EXCEL_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        next_row = 2

    query = row["query"]
    cell = ws.cell(row=next_row, column=1, value=query)
    cell.hyperlink = build_wordstat_url(query)
    cell.font = Font(color="0000FF", underline="single")
    ws.cell(row=next_row, column=2, value=row.get("base_frequency"))
    ws.cell(row=next_row, column=3, value=row.get("exact_frequency"))
    ws.cell(row=next_row, column=4, value=row.get("precise_frequency"))

    for column in ws.columns:
        max_len = 0
        letter = get_column_letter(column[0].column)
        for cell in column:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max_len + 2, 50)

    wb.save(path)


def process_frequencies(
    api: WordstatAPI,
    queries: list[str],
    freq_types: set[str] | None = None,
) -> None:
    selected = set(freq_types) if freq_types else set(DEFAULT_FREQ_TYPES)
    ordered = [t for t in FREQ_TYPES_ORDER if t in selected]
    remaining = list(queries)
    total_start = len(remaining)
    processed = 0
    print(
        f"\nРежим частот: {total_start} запросов в очереди "
        f"({format_freq_types(selected)}; {len(ordered)} API/фразу)..."
    )

    try:
        while remaining:
            seed_query = remaining[0]
            normalized = normalize_query_for_api(seed_query)
            done = processed + 1
            left = len(remaining)
            print(f"\n[{done}/{total_start}] (осталось в файле: {left}) {seed_query}")

            if seed_query != normalized:
                print(f"  нормализация: '{seed_query}' -> '{normalized}'")

            if not normalized:
                print("  ✗ Невалидный запрос после нормализации (пусто)")
                append_invalid_query(seed_query, normalized, "empty_after_normalize")
                remaining.pop(0)
                write_queries_file(remaining)
                continue

            row = {"query": seed_query}
            try:
                values: list[str] = []
                for qtype in ordered:
                    label = FREQ_TYPE_LABELS[qtype]
                    print(f"  {label}...")
                    count = api.get_total_count(format_query(normalized, qtype))
                    row[FREQ_TYPE_ROW_KEYS[qtype]] = count
                    values.append(str(count))
                print(f"  → {' | '.join(values)}")
            except RateLimitError as e:
                print(f"  ✗ {e}")
                continue
            except QuotaExceededStop:
                raise
            except InvalidQueryError as e:
                print(f"  ✗ Невалидный запрос: {e}")
                append_invalid_query(seed_query, normalized, "http_400_invalid_query")
                remaining.pop(0)
                write_queries_file(remaining)
                continue
            except Exception as e:
                print(f"  ✗ Ошибка (seed остаётся в очереди): {e}")
                remaining.append(remaining.pop(0))
                write_queries_file(remaining)
                continue

            append_excel_row(row, REPORT_OUTPUT_FILE)
            remaining.pop(0)
            write_queries_file(remaining)
            processed += 1
            print(f"  ✓ сохранено в Excel, удалено из queries.txt (осталось {len(remaining)})")

    except QuotaExceededStop as e:
        write_queries_file(remaining)
        print(f"\n✗ {e}")
        print(f"  Прогресс сохранён. Осталось в {QUERIES_FILE}: {len(remaining)}")
        return

    print(f"\nГотово: обработано {processed}, файл {REPORT_OUTPUT_FILE}")


def process_phrases(api: WordstatAPI, queries: list[str]) -> None:
    remaining = list(queries)
    total_start = len(remaining)
    print(f"\nРежим сбора фраз «Популярные»: {total_start} seed в очереди...")

    stop_words = read_lines(STOP_WORDS_FILE)
    if stop_words:
        print(f"✓ Загружено минус-слов: {len(stop_words)}")
    else:
        print("ℹ️  Стоп-слова не используются")

    existing = read_lines(WORDS_OUTPUT_FILE)
    seen = {p.lower() for p in existing}
    print(f"✓ Уже в {WORDS_OUTPUT_FILE}: {len(existing)} фраз")

    processed = 0
    try:
        while remaining:
            seed_query = remaining[0]
            normalized = normalize_query_for_api(seed_query)
            done = processed + 1
            left = len(remaining)
            print(f"\n[{done}/{total_start}] (осталось в файле: {left}) {seed_query}")

            if seed_query != normalized:
                print(f"  нормализация: '{seed_query}' -> '{normalized}'")

            if not normalized:
                print("  ✗ Невалидный запрос после нормализации (пусто)")
                append_invalid_query(seed_query, normalized, "empty_after_normalize")
                remaining.pop(0)
                write_queries_file(remaining)
                continue

            try:
                phrases = api.get_popular_phrases(normalized, num_phrases=MAX_PHRASES)
            except RateLimitError as e:
                print(f"  ✗ {e}")
                continue
            except QuotaExceededStop:
                raise
            except InvalidQueryError as e:
                print(f"  ✗ Невалидный запрос: {e}")
                append_invalid_query(seed_query, normalized, "http_400_invalid_query")
                remaining.pop(0)
                write_queries_file(remaining)
                continue
            except Exception as e:
                print(f"  ✗ Ошибка (seed остаётся в очереди): {e}")
                remaining.append(remaining.pop(0))
                write_queries_file(remaining)
                continue

            phrases = filter_stop_words(phrases, stop_words, quiet=True)
            new_phrases: list[str] = []
            for phrase in phrases:
                key = phrase.lower()
                if key not in seen:
                    seen.add(key)
                    new_phrases.append(phrase)

            appended = append_phrases(new_phrases, WORDS_OUTPUT_FILE)
            remaining.pop(0)
            write_queries_file(remaining)
            processed += 1
            print(
                f"  получено {len(phrases)}, новых +{appended} "
                f"(файл слов: {len(seen)}; осталось seed: {len(remaining)})"
            )

    except QuotaExceededStop as e:
        write_queries_file(remaining)
        print(f"\n✗ {e}")
        print(f"  Прогресс сохранён. Осталось в {QUERIES_FILE}: {len(remaining)}")
        return

    print(f"\nГотово: обработано seed {processed}, файл {WORDS_OUTPUT_FILE}")


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Яндекс Вордстат через Cloud API")
    parser.add_argument(
        "--phrases",
        "--deep",
        dest="phrases",
        action="store_true",
        help="Сбор популярных фраз → output/wordstat_words.txt",
    )
    parser.add_argument(
        "--mode",
        choices=["1", "2", "freq", "phrases"],
        help="1/freq — частоты; 2/phrases — сбор фраз",
    )
    parser.add_argument(
        "--freq-types",
        dest="freq_types",
        default=None,
        help="Типы частот: b/e/p или base,exact,precise (по умолчанию только b)",
    )
    return parser.parse_args(argv)


def select_mode(args: argparse.Namespace) -> str:
    if args.phrases or args.mode in ("2", "phrases"):
        return "phrases"
    if args.mode in ("1", "freq"):
        return "freq"

    if sys.stdin and sys.stdin.isatty():
        print("Выберите режим:")
        print("  1 — частоты → output/wordstat_report.xlsx")
        print("  2 — сбор фраз «Популярные» → output/wordstat_words.txt")
        try:
            choice = input("Режим [1/2]: ").strip()
        except EOFError:
            choice = "1"
        return "phrases" if choice == "2" else "freq"

    print("Нет интерактивного ввода — режим 1 (частоты). Для фраз: --phrases")
    return "freq"


def select_freq_types(args: argparse.Namespace) -> set[str]:
    """Выбор типов частот: CLI --freq-types или интерактивный мультивыбор."""
    if args.freq_types is not None:
        return parse_freq_types(args.freq_types)

    if sys.stdin and sys.stdin.isatty():
        print("Типы частот [b/e/p, Enter=только b]:")
        print("  b — базовая")
        print("  e — точная")
        print("  p — уточнённая")
        try:
            raw = input("Типы: ").strip()
        except EOFError:
            raw = ""
        try:
            return parse_freq_types(raw)
        except ValueError as e:
            print(f"✗ {e}")
            print("  Использую дефолт: только базовая")
            return set(DEFAULT_FREQ_TYPES)

    return set(DEFAULT_FREQ_TYPES)


def main(argv: list[str] | None = None) -> None:
    ensure_dirs()
    args = parse_args(argv)
    mode = select_mode(args)

    print("=== Яндекс Вордстат (API) ===\n")
    print("Режим:", "сбор фраз" if mode == "phrases" else "частоты", "\n")

    try:
        api_key, folder_id = load_credentials()
    except (FileNotFoundError, ValueError) as e:
        print(f"✗ {e}")
        return

    queries = read_lines(QUERIES_FILE)
    if not queries:
        print(f"✗ Нет запросов в {QUERIES_FILE}")
        return

    print(f"✓ Запросов в очереди: {len(queries)}")
    api = WordstatAPI(api_key, folder_id)

    try:
        if mode == "phrases":
            process_phrases(api, queries)
        else:
            try:
                freq_types = select_freq_types(args)
            except ValueError as e:
                print(f"✗ {e}")
                return
            print(f"✓ Типы частот: {format_freq_types(freq_types)}")
            process_frequencies(api, queries, freq_types)
    except KeyboardInterrupt:
        print("\nПрервано пользователем (успешно обработанные seed уже сохранены)")
    except Exception as e:
        print(f"\n✗ {e}")


if __name__ == "__main__":
    main()
